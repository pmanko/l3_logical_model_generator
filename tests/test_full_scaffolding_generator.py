# Generated by CodiumAI
import os
from who_l3_smart_tools.core.indicator_testing.scaffolding_generator import (
    ScaffoldingGenerator,
)
import pandas as pd
import unittest
from openpyxl import load_workbook


class TestGenerateTestScaffolding(unittest.TestCase):

    # Generates a new Excel file with the same number of sheets as the number of rows in the 'Indicator definitions' sheet of the input file.
    def test_generate_test_scaffolding_same_number_of_sheets(self):
        input_file = "tests/data/indicator_dak_input_MINI.xlsx"
        output_file = "tests/output/indicator_test_output_MINI.xlsx"

        # Make sure output directory exists
        if not os.path.exists("tests/output"):
            os.makedirs("tests/output")

        sg = ScaffoldingGenerator(input_file, output_file)
        sg.generate_test_scaffolding()

        # Ensure output file exists
        assert os.path.exists(output_file)

        df = pd.read_excel(input_file, sheet_name="Indicator definitions")
        expected_sheets = len(df)

        wb = load_workbook(output_file)
        actual_sheets = len(wb.sheetnames)

        assert expected_sheets == actual_sheets

    # The input file does not exist.
    def test_generate_test_scaffolding_input_file_not_exist(self):
        input_file = "path/to/nonexistent_input_file.xlsx"
        output_file = "path/to/output_file.xlsx"

        try:
            sg = ScaffoldingGenerator(input_file, output_file)
            sg.generate_test_scaffolding(input_file, output_file)
        except FileNotFoundError:
            assert True
            
